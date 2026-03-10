"""
Purpose:
- Execute `B2BSpecialDealsNikonas1` stored procedure and export to Excel.
- Detect and mark NEW rows by comparing current keys with previous export keys.
- Preserve manual override columns and apply report formatting.
"""

import pyodbc
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Color constants used by header/body formatting rules.
NAVY_BLUE = "003366CC"
LIGHT_BLUE = "0027bac2"
BLUE_1 = "00009BFC"
BLUE_2 = "00489ACF"
BLUE_3 = "0080c3d9"
ORANGE = "00F58C40"
ORANGE_2 = "00E35936"
BLUE = "000000FF"
PURPLE = "00C90AF0"
GREEN_LIGHT = "0090EE90"
GREY = "00D3D3D3"
WEIRD_RED = "00DF6666"

# Read-only credentials for SQL report extraction.
password = r'r3adp%78onlygmg'

# SQL connection string for the reporting DB.
conn_str = f"""DRIVER={{SQL Server}};
SERVER=MG-SERVER002;
DATABASE=soft1;
UID=jobview;
PWD={password};"""
# Stored procedure + export workbook path.
stored_procedure = "B2BSpecialDealsNikonas1"
excel_path = r"\\MG-SERVER002\ShareData\reports\Πρότυπα Αρχεία\B2BSpecialDealsNikonas1.xlsx"


def normalize_key_series(series):
    # Normalize keys so joins/maps are stable across numeric/text representations.
    # Steps:
    # 1) cast to pandas string and strip whitespace
    # 2) standardize empty/null tokens to <NA>
    # 3) convert numerics to Int64-backed strings
    # 4) keep non-numeric values unchanged
    raw = series.astype('string').str.strip()
    raw = raw.replace({'': pd.NA, 'nan': pd.NA, 'None': pd.NA, '<NA>': pd.NA})
    numeric = pd.to_numeric(raw, errors='coerce')
    numeric_as_str = numeric.astype('Int64').astype('string')
    return numeric_as_str.where(numeric.notna(), raw)

# Capture previous export keys before refreshing data, used to flag NEW rows.
existing_column_a_values = set()
try:
    # If file/sheet exists, read current IDs from column A.
    if os.path.exists(excel_path):
        df_existing = pd.read_excel(excel_path, sheet_name="B2BSpecialDealsNikonas1", usecols=[0])
        existing_column_a_values = set(normalize_key_series(df_existing.iloc[:, 0]).dropna().tolist())
        print(f"Existing column A values captured: {len(existing_column_a_values)} rows")
except Exception as e:
    # Non-fatal: we can still export; NEW detection may be less accurate.
    print(f"Warning: Could not read existing Excel file: {e}")

new_column_a_values = set()  # Will store rows that are new

try:
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        # Execute report stored procedure.
        cursor.execute(f"EXEC {stored_procedure}")
        # Load schema + rows from SQL cursor.
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        # Convert resultset to DataFrame.
        df = pd.DataFrame.from_records(rows, columns=columns)
        print(f"Rows fetched: {len(df)}")
        
        # Identify new rows (keys present now but not in previous export).
        new_column_a_values = set(normalize_key_series(df.iloc[:, 0]).dropna().tolist()) - existing_column_a_values
        print(f"New column A values found: {len(new_column_a_values)} rows")
        
        # Find STATUS column if it exists.
        col_y_name = None
        for col in df.columns:
            if str(col).upper() == 'STATUS':
                col_y_name = str(col)
                break

        col_a = df.columns[0]
        current_key_series = normalize_key_series(df[col_a])
        # Build NEW flags using normalized keys so type mismatches do not hide new rows.
        new_flags = current_key_series.apply(
            lambda k: "NEW" if pd.notna(k) and k in new_column_a_values else None
        )

        # Always write NEW values to column Z (index 25), which is the reporting column.
        col_z_idx = 25
        # Ensure DataFrame has enough columns to address column Z.
        if len(df.columns) <= col_z_idx:
            while len(df.columns) <= col_z_idx:
                df.insert(len(df.columns), f"EXTRA_{len(df.columns)+1}", None)

        col_z_name = df.columns[col_z_idx]
        df[col_z_name] = new_flags

        # Keep STATUS aligned too, when it exists in the dataset.
        if col_y_name is not None:
            df[col_y_name] = new_flags

        print(f"Column Z populated with 'NEW' for {int(new_flags.eq('NEW').sum())} new rows")
        # Ensure output directory exists before writing Excel file.
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
        # Read previous export columns A/X/Y used for preservation and overrides.
        df_av = pd.read_excel(excel_path, sheet_name="B2BSpecialDealsNikonas1", usecols=[0, 23, 24])
        print(f"Rows read from existing Excel for column 23 preservation: {len(df_av)} rows")

        # Build mapping tables used to preserve manual values and overrides.
        df_av_work = df_av.copy()
        df_av_work['_key'] = normalize_key_series(df_av_work.iloc[:, 0])

        existing_pairs = df_av_work[df_av_work['_key'].notna()]
        # Mapping: key -> existing col 23 value.
        existing_col23_mapping = dict(zip(existing_pairs['_key'], existing_pairs.iloc[:, 1]))

        # Mapping: key -> override col 24 value (only where provided).
        override_pairs = existing_pairs[existing_pairs.iloc[:, 2].notna()]
        override_col24_mapping = dict(zip(override_pairs['_key'], override_pairs.iloc[:, 2]))

        col_a_idx = 0
        col_23_idx = 23
        if len(df.columns) > col_23_idx:
            col_a = df.columns[col_a_idx]
            col_23 = df.columns[col_23_idx]

            # Keep existing column 23 values and only replace them when column 24 has a new override.
            current_key_series = normalize_key_series(df[col_a])
            df[col_23] = current_key_series.map(existing_col23_mapping)
            df[col_23] = current_key_series.map(override_col24_mapping).combine_first(df[col_23])
            print("Column 23 preserved from existing Excel and updated only where column 24 has values")
        
        col_x_idx = 24
        if len(df.columns) > col_x_idx:
            col_x = df.columns[col_x_idx]
            # Clear temporary override column after values were merged.
            df[col_x] = None
            print("Column 24 cleared")
        
        int_columns = {4: 'E', 5: 'F', 6: 'G', 7: 'H', 18: 'S', 19: 'T', 20: 'U', 21: 'V'}
        # Coerce selected fields to nullable integers for clean Excel output.
        for idx, col_name in int_columns.items():
            if len(df.columns) > idx:
                col = df.columns[idx]
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')
        
        # Export refreshed data before the styling pass.
        df.to_excel(excel_path, index=False, sheet_name="B2BSpecialDealsNikonas1")
        print(f"Excel exported to:\n{excel_path}")
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Build style objects once and reuse in loops for performance/readability.
        navy_blue_fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
        light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        orange_2_fill = PatternFill(start_color=ORANGE_2, end_color=ORANGE_2, fill_type="solid")
        purple_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
        blue_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        blue_1_fill = PatternFill(start_color=BLUE_1, end_color=BLUE_1, fill_type="solid")
        blue_2_fill = PatternFill(start_color=BLUE_2, end_color=BLUE_2, fill_type="solid")
        blue_3_fill = PatternFill(start_color=BLUE_3, end_color=BLUE_3, fill_type="solid")
        red_font = Font(color="00FF0000")
        green_light_fill = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
        weird_red_fill = PatternFill(start_color=WEIRD_RED, end_color=WEIRD_RED, fill_type="solid")
        
        bold_font = Font(bold=True)
        
        # Header colors by logical column groups.
        for col in ['A', 'B', 'C', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
            ws[f'{col}1'].fill = navy_blue_fill
        
        for col in ['X', 'Y']:
            ws[f'{col}1'].fill = purple_fill
        
        for col in ['E', 'F', 'G']:
            for row in range(2, ws.max_row + 1):
                ws[f'{col}{row}'].fill = orange_fill
        
        # X/Y are manual-action columns: orange background + border when populated.
        for col in ['X', 'Y']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                cell.fill = orange_2_fill
                if cell.value is None or str(cell.value).strip() == '':
                    continue
                cell.border = Border(top=Side(border_style="thin", color="000000"),
                                     left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                cell.font = bold_font
        
        for row in range(2, ws.max_row + 1):
                ws[f'{'I'}{row}'].fill = blue_1_fill
        
        # Column J is highlighted with bold red-ish fill for emphasis.
        for row in range(2, ws.max_row + 1):
                ws[f'{'J'}{row}'].fill = weird_red_fill
                ws[f'{'J'}{row}'].font = bold_font
                
        for col in ['K', 'L', 'M', 'N']:
            for row in range(2, ws.max_row + 1):
                ws[f'{col}{row}'].fill = blue_2_fill
        
        for row in range(2, ws.max_row + 1):
            ws[f'{'H'}{row}'].fill = blue_3_fill
        
        for col in ['S', 'T', 'U', 'V', 'W']:
            for row in range(1, ws.max_row + 1):
                ws[f'{col}{row}'].font = red_font
        
        # Apply blue fill and white font to column Y for NEW rows
        blue_fill_for_new = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        white_font = Font(color="00FFFFFF")
        for row in range(2, ws.max_row + 1):
            if ws[f'Z{row}'].value == "NEW":
                ws[f'Z{row}'].fill = blue_fill_for_new
                ws[f'Z{row}'].font = white_font
        
        ws['E1'].fill = green_light_fill
        ws['F1'].fill = light_blue_fill
        ws['G1'].fill = light_blue_fill
        
        for row in range(2, ws.max_row + 1):
            if ws[f'A{row}'].value is None or str(ws[f'A{row}'].value).strip() == '':
                # Grey-out separator/header rows where column A is empty.
                grey_fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row, col_idx).fill = grey_fill
            else:
                # Keep ID rows highlighted in light blue.
                ws[f'A{row}'].fill = light_blue_fill
        
        # Expand comma-separated supplier list into additional columns.
        max_tokens = 0
        token_data = []
        
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row, 27).value
            if cell_value:
                tokens = [t.strip() for t in str(cell_value).split(",")]
                tokens = [t for t in tokens if t]
                token_data.append((row, tokens))
                if len(tokens) > max_tokens:
                    max_tokens = len(tokens)
        
        cnt=1
        
        from openpyxl.utils import get_column_letter
        for i in range(max_tokens):
            col_num = 28 + i
            col_letter = get_column_letter(col_num)
            # Dynamic supplier headers: Προμηθευτής 2, 3, 4, ...
            ws[f'{col_letter}1'] = f"Προμηθευτής {i+cnt+1}"
        
        for row, tokens in token_data:
            for i, token in enumerate(tokens):
                col_num = 28 + i
                ws.cell(row, col_num).value = token
        
        # ws.column_dimensions['A'].width = 25
        # ws.column_dimensions['B'].width = 25
        # ws.column_dimensions['C'].width = 25
        # ws.column_dimensions['D'].width = 25
        # ws.column_dimensions['E'].width = 25
        # ws.column_dimensions['F'].width = 25
        # ws.column_dimensions['G'].width = 25
        # ws.column_dimensions['H'].width = 25
        # ws.column_dimensions['I'].width = 25
        # ws.column_dimensions['J'].width = 25
        # ws.column_dimensions['K'].width = 25
        # ws.column_dimensions['L'].width = 25
        # ws.column_dimensions['M'].width = 25
        # ws.column_dimensions['N'].width = 25
        # ws.column_dimensions['O'].width = 25
        # ws.column_dimensions['P'].width = 25
        # ws.column_dimensions['Q'].width = 25
        # ws.column_dimensions['R'].width = 25
        # ws.column_dimensions['S'].width = 25
        # ws.column_dimensions['T'].width = 25
        # ws.column_dimensions['U'].width = 25
        # ws.column_dimensions['V'].width = 25
        # ws.column_dimensions['W'].width = 25
        # ws.column_dimensions['X'].width = 25
        # ws.column_dimensions['Y'].width = 25
        # ws.column_dimensions['Z'].width = 25
        # ws.column_dimensions['AA'].width = 25

        ws.auto_filter.ref = ws.dimensions
        
        # Slight zoom-out for easier on-screen scanning.
        ws.sheet_view.zoomScale = 85

        print(f"Unpacked {max_tokens} columns from comma-separated values")
        
        wb.save(excel_path)
        print("Color formatting applied")
        
except Exception as e:
    print("ERROR:")
    print(e)
    
#pyinstaller --onefile B2BSpecialDealsNikonas1.py # \\MG-SERVER002\ShareData\reports\exe