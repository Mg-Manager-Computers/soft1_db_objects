"""
Purpose:
- Execute the `DSV` stored procedure against SQL Server.
- Export results to the DSV Excel template.
- Apply layout and conditional formatting used by the reporting team.
"""

import pyodbc
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Color palette used by report formatting rules.
DARK_GREY = "00656363"
LIGHT_GREY = "00E1E1E1"
LIGHT_BLUE = "0053DEEE"
DEEP_BLUE = "0030398A"
BLUE = "003A98F5"
GREEN = "0007D407"
LIGHT_GREEN = "0090EE90"
WEIRD_GREEN = "004C9900"
VERY_LIGHT_ORANGE = "FFCC87"
LIGHT_ORANGE = "00F9C04D"
ORANGE = "00F58C40"
PURPLE = "00DA42F9"

# Read-only credentials for report extraction.
password = r'r3adp%78onlygmg'

# Explanatory notes rendered on row 1 for key report columns.
TEXT_Q = 'Περιλαμβάνει το συνολο των πωλησεων\n(Skroutz hub included)'
TEXT_S = 'Υπολογίζεται με βάση μόνο τις πωλήσεις MG\n(Skroutz Sales excluded)'
TEXT_T = 'Aφορά Διακίνηση από DSV σε Παπαδά\nκαι έχει αφαιρεθεί από το suggestion'
TEXT_U = 'Υπολογίζεται με βάση μόνο\nτις πωλήσεις Skroutz'
TEXT_V = 'Aφορά Διακίνηση από οποιοδήποτε\nΑποθ.Χώρο προς Skroutz\nκαι θα έχει αφαιρεθεί από το suggestion'
TEXT_W = 'Όσα τεμάχια είναι σε εικονική αναμονή\nαλλά όχι σε διακίνηση'

# SQL Server connection details.
conn_str = f"""DRIVER={{SQL Server}};
SERVER=MG-SERVER002;
DATABASE=soft1;
UID=jobview;
PWD={password};"""
# Target stored procedure and report file.
stored_procedure = "DSV"
excel_path = r"\\MG-SERVER002\ShareData\reports\Πρότυπα Αρχεία\DSV.xlsx"

# Main ETL + formatting pipeline.
try:
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        # Run the stored procedure and load the resultset into a DataFrame.
        cursor.execute(f"EXEC {stored_procedure}")
        # Read output schema and rows from SQL cursor.
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        # Convert SQL rows into a DataFrame for tabular export.
        df = pd.DataFrame.from_records(rows, columns=columns)
        print(f"Rows fetched: {len(df)}")
        
        # Open the existing workbook template.
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Clear all data from the worksheet (keep the sheet, remove all rows except header if any)
        max_row = int(ws.max_row) if ws.max_row else 0
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        
        # Write the dataframe data starting from row 1
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Get the number of rows after writing data (for styling loops)
        data_rows = int(ws.max_row) if ws.max_row else 0
        
        # Now insert a new row at position 1 for the TEXT values (this pushes data to row 2)
        ws.insert_rows(1)
        
        # Increase top row height so multi-line explanatory text is visible.
        ws.row_dimensions[1].height = 75
        
        # Widen columns containing explanatory notes.
        ws.column_dimensions['Q'].width = 25
        ws.column_dimensions['S'].width = 25
        ws.column_dimensions['T'].width = 25
        ws.column_dimensions['U'].width = 25
        ws.column_dimensions['V'].width = 25
        ws.column_dimensions['W'].width = 25
        
        ws['Q1'] = TEXT_Q
        ws['Q1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws['S1'] = TEXT_S
        ws['S1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws['T1'] = TEXT_T
        ws['T1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws['U1'] = TEXT_U
        ws['U1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws['V1'] = TEXT_V
        ws['V1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws['W1'] = TEXT_W
        ws['W1'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Update data_rows to reflect the inserted row
        data_rows = int(ws.max_row) if ws.max_row else 0
        
        # Add filter to row 2 (the header row now, after inserting TEXT row at position 1)
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f'A2:{last_col}{ws.max_row}'
        
        # Initialize fills/fonts used by the style rules below.
        dark_grey_fill = PatternFill(start_color=DARK_GREY, end_color=DARK_GREY, fill_type="solid")
        light_grey_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
        light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        very_light_orange_fill = PatternFill(start_color=VERY_LIGHT_ORANGE, end_color=VERY_LIGHT_ORANGE, fill_type="solid")
        light_orange_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
        orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        purple_fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
        blue_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        deep_blue_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
        green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        light_green_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
        weird_green_fill = PatternFill(start_color=WEIRD_GREEN, end_color=WEIRD_GREEN, fill_type="solid")
        
        red_font = Font(color="00FF0000", bold=True)
        light_grey_font = Font(color=LIGHT_GREY)
        dark_grey_font = Font(color=DARK_GREY)
        bold_font = Font(bold=True)
        
        # Special header accent for column L.
        ws[f'L2'].fill = weird_green_fill 
        ws[f'L2'].font = dark_grey_font
        
        # Apply business color coding for stock and demand metric columns.
        for col in ['G', 'H', 'I']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = green_fill
                ws[f'{col}{row}'].font = bold_font
        
        for col in ['J', 'K', 'L']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = light_orange_fill
        
        # Column W is highlighted red when positive to draw attention.
        for col in ['P', 'W']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = light_blue_fill
                
                if col == 'W':
                    cell_value = ws[f'{col}{row}'].value
                    try:
                        numeric_value = float(cell_value) if cell_value is not None else 0
                        if numeric_value > 0:
                            ws[f'{col}{row}'].font = red_font
                        else:
                            ws[f'{col}{row}'].font = bold_font
                    except (ValueError, TypeError):
                        ws[f'{col}{row}'].font = bold_font
        
        for row in range(3, data_rows + 1):     
            ws[f'Q{row}'].fill = very_light_orange_fill
            
        # Suggestion columns are highlighted and bolded for quick review.
        for col in ['R', 'S', 'U']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = orange_fill
                ws[f'{col}{row}'].font = bold_font
        
        # Transfer-related columns use blue background.
        for col in ['T', 'V']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = blue_fill
                ws[f'{col}{row}'].font = bold_font
        
        
        # Header color groups.
        for col in ['A', 'B', 'F', 'G', 'H', 'P', 'Q', 'R', 'S', 'U']:
            ws[f'{col}2'].fill = dark_grey_fill
            ws[f'{col}2'].font = light_grey_font
            
        for col in ['C', 'D', 'E', 'V']:
            ws[f'{col}2'].fill = light_blue_fill
            ws[f'{col}2'].font = light_grey_font
            
        for col in ['I', 'J', 'K']:
            ws[f'{col}2'].fill = deep_blue_fill
            ws[f'{col}2'].font = light_grey_font
            
        for col in ['M', 'N', 'O']:
            ws[f'{col}2'].fill = blue_fill
            ws[f'{col}2'].font = dark_grey_font
        
        for col in ['M', 'N', 'O']:
            for row in range(3, data_rows + 1):
                ws[f'{col}{row}'].fill = light_green_fill
        
        # Emphasize key header columns T and W.
        for col in ['T', 'W']:
            ws[f'{col}2'].fill = purple_fill
            ws[f'{col}2'].font = light_grey_font
        
        # Slight zoom-out for readability on dense reports.
        ws.sheet_view.zoomScale = 85
        
        # Save the workbook
        wb.save(excel_path)
        
except Exception as e:
    print("ERROR:")
    print(e)
    
#pyinstaller --onefile DSV.py # \\MG-SERVER002\ShareData\reports\exe